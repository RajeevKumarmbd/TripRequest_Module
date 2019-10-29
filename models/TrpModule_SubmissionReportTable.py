from odoo import models, fields
import io
import base64
import openpyxl
from datetime import timedelta
from os import path
from openpyxl.styles import PatternFill
from tempfile import TemporaryFile


class TrpModuleSubmissionReportTable(models.Model):
    _name = 'trpmodule.submissionreporttable'

    from_dt = fields.Date('Date', required=True)
    name = fields.Char('Submission Report')
    upload_file = fields.Binary(string="Upload Submission Report")
    uploadedfilename = fields.Char('File Name', size=256, default='Select Submission Report')
    remarks = fields.Text('Remark')
    excel_da = fields.Text('Excel Data')

    def get_fcm_data(self, data, context=None):
        fcmbookinvoice_obj = self.env['trpmodule.fcmbookeddetailstable']
        fcmbookinvoice_ids = fcmbookinvoice_obj.search(
            ['&', ('travel_date', '>=', self.from_dt), ('fcm_status', '=', 'True')])
        self.remarks = 'Start'
        for thisheader_ids in fcmbookinvoice_ids:
            self.remarks = self.remarks + '{ invoice No. : ' + str(
                thisheader_ids.invoice_no) + ', Ticket No. - ' + thisheader_ids.name + ', TRIP ID - ' + thisheader_ids.trip_id + '}'

    # def get_excel_data(self, data, context=None):
    #     submission_obj = self.env['trpmodule.submissiontable']
    #     submission_ids = submission_obj.search(['&', ('name', '=', 'U0680721'), ('tkt_no', '=', 'FP6ZNB-1')])
    #     self.excel_da = 'Show'
    #     for thissubmission_ids in submission_ids:
    #         self.excel_da = self.excel_da + '{ invoice No. : ' + str(
    #             thissubmission_ids.name) + ', Ticket No. - ' + thissubmission_ids.tkt_no + '} '

    def upload_excel(self, data, context=None):
        file = base64.decodestring(self.upload_file)
        excel_fileobj = TemporaryFile('wb+')
        excel_fileobj.write(file)
        excel_fileobj.seek(0)
        wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
        wb.active = 0
        worksheet = wb.active
        itempos = 5
        maxmimum_row = 500
        self.remarks = ''
        for i in range(1, maxmimum_row + 1, 1):
            submissioonreport_obj = self.env['trpmodule.submissiontable']
            self.ensure_one()
            advice_no = worksheet.cell(row=i, column=1)
            sr_no = worksheet.cell(row=i, column=2)
            message = worksheet.cell(row=i, column=3)
            agents_irno = worksheet.cell(row=i, column=4)
            agency_irno = worksheet.cell(row=i, column=5)
            upload_batchno = worksheet.cell(row=i, column=6)
            ho_non_ho = worksheet.cell(row=i, column=7)
            domestic_int = worksheet.cell(row=i, column=8)
            project_code = worksheet.cell(row=i, column=9)
            tran_type = worksheet.cell(row=i, column=10)
            isgec_emp_code = worksheet.cell(row=itempos, column=11)
            isgec_gstin = worksheet.cell(row=i, column=12)
            supplier_name = worksheet.cell(row=i, column=13)
            supplier_gstin = worksheet.cell(row=i, column=14)
            name = worksheet.cell(row=itempos, column=15)
            # invoice_date = worksheet.cell(row=i, column=16)
            pax_name = worksheet.cell(row=i, column=17)
            sector = worksheet.cell(row=i, column=18)
            airline_name = worksheet.cell(row=i, column=19)
            # travel_date = worksheet.cell(row=i, column=20)
            tkt_no = worksheet.cell(row=i, column=21)
            booked = worksheet.cell(row=i, column=22)
            service_charge = worksheet.cell(row=i, column=23)
            igst_amt = worksheet.cell(row=i, column=24)
            sgst_amt = worksheet.cell(row=i, column=25)
            cgst_amt = worksheet.cell(row=i, column=26)
            total_service_amt = worksheet.cell(row=i, column=27)
            total_ticket_amt = worksheet.cell(row=i, column=28) or 5555
            total_amt = worksheet.cell(row=i, column=29)
            ref_no = worksheet.cell(row=i, column=30)
            state_code = worksheet.cell(row=i, column=31)
            rcm_applicable = worksheet.cell(row=i, column=32)
            gst_tax_category = worksheet.cell(row=i, column=33)
            airline_gstin = worksheet.cell(row=i, column=34)
            tkt_no_2 = worksheet.cell(row=i, column=35)
            booked_on = worksheet.cell(row=i, column=36)
            airline_name_2 = worksheet.cell(row=i, column=37)
            basic_fare = worksheet.cell(row=i, column=38)
            yq_tax_amt = worksheet.cell(row=i, column=39)
            isgt_amt_2 = worksheet.cell(row=i, column=40)
            sgst_amt_2 = worksheet.cell(row=i, column=41)
            cgst_amt_2 = worksheet.cell(row=i, column=42)
            all_other_taxes_amt = worksheet.cell(row=i, column=43)
            total_ticket_amt_2 = worksheet.cell(row=i, column=44)
            bta_btc = worksheet.cell(row=i, column=45)
            submission_period = worksheet.cell(row=i, column=46)
            if name.value is not None:
                submissioonreport_obj.create(
                    {'name': name.value,
                     'advice_no': advice_no.value,
                     'sr_no': sr_no.value,
                     'message': message.value,
                     'agents_irno': agents_irno.value,
                     'agency_irno': agency_irno.value,
                     'upload_batchno': upload_batchno.value,
                     'ho_non_ho': ho_non_ho.value,
                     'domestic_international': domestic_int.value,
                     'project_code': project_code.value,
                     'tran_type': tran_type.value,
                     'isgec_emp_code': isgec_emp_code.value,
                     'isgec_gstin': isgec_gstin.value,
                     'supplier_name': supplier_name.value,
                     'supplier_gstin': supplier_gstin.value,
                     'invoice_date': '2019/10/01',
                     'pax_name': pax_name.value,
                     'sector': sector.value,
                     'airline_name': airline_name.value,
                     'travel_date': '2019/10/01',
                     'tkt_no': tkt_no.value,
                     'booked': booked.value,
                     'service_charge': service_charge.value,
                     'igst_amt': '0.00',
                     'sgst_amt': '0.00',
                     'cgst_amt': '0.00',
                     'total_service_amt': '0.00',
                     'total_ticket_amt': '0.00',
                     'total_amt': '0.00',
                     'ref_no': ref_no.value,
                     'state_code': state_code.value,
                     'rcm_applicable': rcm_applicable.value,
                     'gst_tax_category': gst_tax_category.value,
                     'airline_gstin': airline_gstin.value,
                     'tkt_no_2': tkt_no_2.value,
                     'booked_on': booked_on.value,
                     'airline_name_2': airline_name_2.value,
                     'basic_fare': basic_fare.value,
                     'yq_tax_amt': '0.00',
                     'isgt_amt_2': '0.00',
                     'sgst_amt_2': '0.00',
                     'cgst_amt_2': '0.00',
                     'all_other_taxes_amt': '0.00',
                     'total_ticket_amt_2': '0.00',
                     'bta_btc': bta_btc.value,
                     'submission_period': submission_period.value})
                itempos = itempos + 1
        self.remarks = 'Record Created Successfully'
