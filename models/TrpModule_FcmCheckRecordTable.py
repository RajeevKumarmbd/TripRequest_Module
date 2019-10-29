from os import path
from AptUrl.Helpers import _
from odoo import models, fields
import io
import base64
import openpyxl
from tempfile import TemporaryFile


class TrpModuleFcmCheckRecordTable(models.Model):
    _name = 'trpmodule.fcmcheckrecordtable'

    from_dt = fields.Date('Date', required=True)
    name = fields.Char('Submission Report')
    upload_file = fields.Binary(string="Upload FCM invoice Report")
    uploadedfilename = fields.Char('File Name', size=256, default='Select Submission Report')
    remarks = fields.Text('Remarks')

    download_file = fields.Binary(string="Download")
    downloadedfilename = fields.Char('File Name', size=256, default='Download FCM List.xlsx')

    def get_excel_data(self, data, context=None):
        file = base64.decodestring(self.upload_file)
        excel_fileobj = TemporaryFile('wb+')
        excel_fileobj.write(file)
        excel_fileobj.seek(0)
        wb = openpyxl.load_workbook(excel_fileobj, data_only=True)
        wb.active = 2
        worksheet = wb.active
        itempos = 5
        maxmimum_row = 6
        self.excel_da = ''

        #  Getting Record from Excel Booked Table
        a = []
        e = []

        for i in range(2, maxmimum_row + 1, 1):
            tktno = worksheet.cell(row=itempos, column=21)
            invoice_no = worksheet.cell(row=itempos, column=15)
            #  Getting Record Excel Sheet
            if invoice_no.value is not None:
                a.append(invoice_no.value)
                e.append(tktno.value)
            # self.excel_da = self.excel_da + ' ' + str(travel_dt.value) + ','
            self.excel_da = 'Excel Data List : ' + str(a)
            itempos = itempos + 1

            #  Getting Record from  FCM  Table
            fcmbookinvoice_obj = self.env['trpmodule.fcmbookeddetailstable']
            fcmbookinvoice_ids = fcmbookinvoice_obj.search([('travel_date', '>=', self.from_dt)])
            self.remarks = ''
            b = []
            f = []

            for thisheader_ids in fcmbookinvoice_ids:
                b.append(thisheader_ids.invoice_no)
                f.append(thisheader_ids.name)
                # self.remarks = self.remarks + ' ' + str(thisheader_ids.invoice_no) + ','
                # self.remarks = 'FCM Booking Record Matched List : ' + str(b)

            for val in b:
                if val in a:
                    a.remove(val)

            for val in f:
                if val in e:
                    e.remove(val)
            self.remarks = 'Not Matched Record from FCM Data : ' + str(a)

        # src = path.dirname(path.realpath(__file__)) + "/FCMBOOKINGDETAILS.xlsx"
        # wb = openpyxl.load_workbook(src)
        wb.active = 2
        worksheet = wb.active
        # rowpos = 2
        # rowposis = 2
        itempos = 5
        if self.from_dt:
            for invicelist in a:
                setcol1 = worksheet.cell(row=itempos, column=49)
                setcol1.value = invicelist or ''
                itempos = itempos + 1

            # for tiktno in e:
            #     setcol1 = worksheet.cell(row=itempos, column=49)
            #     setcol1.value = tiktno or ''
            #     itempos = itempos + 1

        wb.active = 2
        worksheet = wb.active
        fp = io.BytesIO()
        wb.save(fp)
        out = base64.encodestring(fp.getvalue())
        self.download_file = out
